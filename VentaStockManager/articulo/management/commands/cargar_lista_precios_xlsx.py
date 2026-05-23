"""
Carga lista de precios desde un .xlsx directo (no texto). Mejor que
el comando de texto: usa el formato `font.bold` de la celda para
distinguir headers de categoría de filas de artículo — cero adivinanza.

Estructura del Excel esperada (formato de "Golosinas Insa", mayo 2026):
  - Col A: nombre del artículo (o nombre de categoría si la fila está bold)
  - Col B: código (etiqueta humana — A1, B23, M00, etc.)
  - Col D: precio mayorista (con $)
  - Las categorías aparecen como filas con **negrita** en col A, sin
    valor en cols B/D.

Mapeo categoría → rubro:
  El Excel SOLO tiene categorías (1 nivel). El operador quiere agrupar
  categorías en rubros (Golosinas, Bebidas, etc.) para listas de precios
  más eficientes. Como el Excel no tiene esa info, este comando lleva
  un mapeo razonable hardcoded abajo (`DEFAULT_CATEGORIA_A_RUBRO`).

  - Si una categoría del Excel está en el mapeo, se asigna a ese rubro.
  - Si no, se asigna al rubro "Otros".
  - Podés pasar --rubros-mapeo path/to/mapeo.json para overrider.

Idempotencia:
  - Rubro: get_or_create por nombre (case-insensitive).
  - Categoría: get_or_create por nombre (case-insensitive). Si ya existe,
    se le pone el rubro encontrado (si no tenía ya uno).
  - Articulo: match por (codigo) si está, sino por nombre + categoria.
    Si --update-precios, actualiza precio_mayorista.

Uso:
  python manage.py cargar_lista_precios_xlsx /path/a/articulos.xlsx
  python manage.py cargar_lista_precios_xlsx /path/a/articulos.xlsx --dry-run
  python manage.py cargar_lista_precios_xlsx /path/a/articulos.xlsx --update-precios
  python manage.py cargar_lista_precios_xlsx /path/a/articulos.xlsx --hoja articulos
  python manage.py cargar_lista_precios_xlsx /path/a/articulos.xlsx --col-precio E   # si el precio está en otra columna
"""
from __future__ import annotations

import json
import re
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from articulo.models import Articulo, Categoria, Rubro


VENCIMIENTO_SENTINEL = date(2099, 12, 31)


# Mapeo categoría → rubro pensado para "Golosinas Insa". Si la
# categoría no está acá, va al rubro fallback (`--rubro-default Otros`
# por default).
#
# Las KEYS son los nombres EXACTOS como aparecen en el Excel (ya
# normalizados — limpieza de `*`, `_`, etc.). Si en futuras versiones
# del Excel cambian, hay que ajustar acá.
DEFAULT_CATEGORIA_A_RUBRO: dict[str, str] = {
    # Golosinas: dulces clásicos de mostrador
    'MASTICABLE(caramelo)': 'Golosinas',
    'GOMITASS': 'Golosinas',
    'PASTILLAS': 'Golosinas',
    'ALFAJORES': 'Golosinas',
    'CHOCOLATES': 'Golosinas',
    'CHUPETINES': 'Golosinas',
    'CHICLES': 'Golosinas',

    # Galletas y Snacks: paquetes saladitos / dulces no-chocolate
    'GALLETAS': 'Galletas y Snacks',
    'SNACK': 'Galletas y Snacks',

    # Bebidas: gaseosas, jugos en polvo, aguas, alcoholes
    'Bebidas': 'Bebidas',
    'JUGOS TANG': 'Bebidas',
    'JUGOS CLIGHT': 'Bebidas',
    'JUGOS RINDE 2': 'Bebidas',
    'JUGOS NOEL': 'Bebidas',
    'JUGOS JA!': 'Bebidas',
    'BEBIDAS ALCOHOLICAS': 'Bebidas',

    # Almacén: comida no perecedera
    'ALMACÉN': 'Almacén',
    'CONDIMENTO Y SABORES': 'Almacén',

    # Limpieza e Higiene: productos del hogar + cuidado personal
    'LIMPIEZA': 'Limpieza e Higiene',
    'HIGIENE PERSONAL': 'Limpieza e Higiene',

    # Helados: todo lo del freezer + insumos
    'HELADOS': 'Helados',
    'Línea tasitas': 'Helados',
    'Línea postres': 'Helados',
    'Línea familiar': 'Helados',
    'Tarros de 10 litros de agua': 'Helados',
    'Tarros de 10 litros sabores comunes': 'Helados',
    'Tarros de 10 litros sabores especiales': 'Helados',
    'Tarros de 10 litros SÚPER sabores': 'Helados',
    'INSUMOS PARA HELADERIA': 'Helados',

    # Salud: analgésicos OTC
    'ANALGÉSICOS 💊': 'Salud',

    # Tabaco (rubro aparte por regulación y márgenes distintos)
    'CIGARRILLOS.': 'Tabaco',

    # Estacionales (productos que se venden solo en ciertas épocas)
    'PRODUCTOS NAVIDEnOS': 'Estacional',
    'PIROTECNIA': 'Estacional',

    # Otros: cajón de sastre
    'VARIOS': 'Otros',
    'INPORTADOS': 'Otros',
    'EXTRAS': 'Otros',
}


# Colores para los rubros — Material design palette. Hacen las cards
# del editor de lista de precios más distinguibles a primera vista.
RUBRO_COLORES: dict[str, str] = {
    'Golosinas': '#EC4899',          # rosa
    'Galletas y Snacks': '#F59E0B',  # ámbar
    'Bebidas': '#3B82F6',            # azul
    'Almacén': '#10B981',            # verde
    'Limpieza e Higiene': '#06B6D4', # cyan
    'Helados': '#8B5CF6',            # violeta
    'Salud': '#EF4444',              # rojo
    'Tabaco': '#78716C',             # gris cálido
    'Estacional': '#F97316',         # naranja
    'Otros': '#9CA3AF',              # gris neutro
}


def _clean_categoria(s: str) -> str:
    """Igual que el comando de texto — saca markdown / símbolos sueltos."""
    s = re.sub(r'[*_]', '', s).strip()
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def _clean_nombre_articulo(s: str) -> str:
    """Saca bullets/dots al inicio + colapsa espacios."""
    s = s.strip()
    s = re.sub(r'^[•\.\*\s]+', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def _parse_precio(raw) -> Decimal | None:
    """
    Parsea precio del Excel. Acepta:
      - número int/float ('4800') → Decimal('4800')
      - string '$4800', '$ 4500', 'S1200' (typo), '4500,50'
    Rechaza: 0, negativos, '001', '0000', > 10M.
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        d = Decimal(str(raw))
    else:
        s = str(raw).strip()
        s = s.lstrip('$Ss').strip()
        s = s.replace(' ', '').replace(',', '.')
        if not re.match(r'^\d+(\.\d+)?$', s):
            return None
        try:
            d = Decimal(s)
        except InvalidOperation:
            return None
    if d <= 0 or d > Decimal('10000000'):
        return None
    return d


class Command(BaseCommand):
    help = (
        'Carga categorías y artículos desde un .xlsx, agrupando '
        'categorías en rubros según un mapeo configurable. Idempotente.'
    )

    def add_arguments(self, parser):
        parser.add_argument('path', type=str, help='Path al archivo .xlsx')
        parser.add_argument(
            '--hoja', type=str, default='articulos',
            help='Nombre de la hoja (default: "articulos")',
        )
        parser.add_argument(
            '--col-nombre', type=str, default='A',
            help='Letra de columna del nombre (default: A)',
        )
        parser.add_argument(
            '--col-codigo', type=str, default='B',
            help='Letra de columna del código (default: B)',
        )
        parser.add_argument(
            '--col-precio', type=str, default='D',
            help='Letra de columna del precio (default: D, que en este '
                 'Excel tiene el precio MAYORISTA aunque el header diga '
                 '"minorista")',
        )
        parser.add_argument(
            '--rubros-mapeo', type=str, default=None,
            help='Path a un JSON con override de mapeo categoría→rubro. '
                 'Formato: {"NOMBRE_CATEGORIA": "Nombre Rubro", ...}',
        )
        parser.add_argument(
            '--rubro-default', type=str, default='Otros',
            help='Rubro para categorías sin mapeo (default: "Otros")',
        )
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument(
            '--update-precios', action='store_true',
            help='Si un artículo ya existe, actualizar precio_mayorista.',
        )
        parser.add_argument(
            '--minorista-igual-mayorista', action='store_true',
            help='Al CREAR un artículo, copiar mayorista a minorista también.',
        )
        parser.add_argument('--verbose', action='store_true')

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise CommandError(
                f'openpyxl no instalado: {e}. Agregalo a requirements.txt '
                f'o ejecutá `pip install openpyxl`.'
            )

        path = Path(opts['path'])
        if not path.exists():
            raise CommandError(f'Archivo no encontrado: {path}')

        # Cargar workbook SIN data_only para tener acceso a font.bold
        # (data_only=True devolvería resultado de fórmulas pero pierde
        # algo del formato; no usamos fórmulas acá así que es indiferente).
        wb = load_workbook(path, data_only=True)
        if opts['hoja'] not in wb.sheetnames:
            raise CommandError(
                f'Hoja "{opts["hoja"]}" no existe. Disponibles: '
                f'{wb.sheetnames}'
            )
        ws = wb[opts['hoja']]

        col_nombre = ws[f"{opts['col_nombre']}1"].column
        col_codigo = ws[f"{opts['col_codigo']}1"].column
        col_precio = ws[f"{opts['col_precio']}1"].column

        # Cargar mapeo categoría → rubro (default + override del JSON si hay)
        mapeo: dict[str, str] = dict(DEFAULT_CATEGORIA_A_RUBRO)
        if opts['rubros_mapeo']:
            with Path(opts['rubros_mapeo']).open('r', encoding='utf-8') as f:
                mapeo.update(json.load(f))
        rubro_default = opts['rubro_default']

        dry_run = opts['dry_run']
        update_precios = opts['update_precios']
        minorista_igual = opts['minorista_igual_mayorista']
        verbose = opts['verbose']

        stats = {
            'rubros_creados': 0,
            'rubros_existentes': 0,
            'categorias_creadas': 0,
            'categorias_existentes': 0,
            'articulos_creados': 0,
            'articulos_actualizados': 0,
            'articulos_sin_cambios': 0,
            'skipped': 0,
        }

        self.stdout.write(self.style.WARNING(
            f'{"DRY RUN — " if dry_run else ""}'
            f'Procesando {ws.max_row} filas de {path.name} (hoja: {opts["hoja"]})\n'
        ))

        # Caches en memoria para no pegarle 1 SELECT por fila.
        cache_rubros: dict[str, Rubro | None] = {}
        cache_categorias: dict[str, Categoria | None] = {}

        def _get_or_create_rubro(nombre: str) -> Rubro | None:
            key = nombre.lower().strip()
            if key in cache_rubros:
                return cache_rubros[key]
            existente = Rubro.objects.filter(nombre__iexact=nombre).first()
            if existente:
                stats['rubros_existentes'] += 1
                cache_rubros[key] = existente
                return existente
            if dry_run:
                cache_rubros[key] = None
                stats['rubros_creados'] += 1
                return None
            color = RUBRO_COLORES.get(nombre, '#9CA3AF')
            nuevo = Rubro.objects.create(nombre=nombre, color=color)
            stats['rubros_creados'] += 1
            cache_rubros[key] = nuevo
            return nuevo

        def _get_or_create_categoria(nombre: str, rubro: Rubro | None) -> Categoria | None:
            key = nombre.lower().strip()
            if key in cache_categorias:
                return cache_categorias[key]
            existente = Categoria.objects.filter(nombre__iexact=nombre).first()
            if existente:
                stats['categorias_existentes'] += 1
                # Si la categoría existía pero NO tenía rubro, ahora se lo
                # asignamos (lo más común: corrí el comando viejo TXT que
                # no manejaba rubros).
                if rubro and not existente.rubro_id and not dry_run:
                    existente.rubro = rubro
                    existente.save(update_fields=['rubro'])
                cache_categorias[key] = existente
                return existente
            if dry_run:
                cache_categorias[key] = None
                stats['categorias_creadas'] += 1
                return None
            nueva = Categoria.objects.create(nombre=nombre, rubro=rubro)
            stats['categorias_creadas'] += 1
            cache_categorias[key] = nueva
            return nueva

        with transaction.atomic():
            categoria_actual: Categoria | None = None

            for row in ws.iter_rows(min_row=2):  # skip headers (R1)
                cell_nombre = row[col_nombre - 1]
                cell_codigo = row[col_codigo - 1] if len(row) >= col_codigo else None
                cell_precio = row[col_precio - 1] if len(row) >= col_precio else None

                val_nombre = cell_nombre.value
                val_codigo = cell_codigo.value if cell_codigo else None
                val_precio = cell_precio.value if cell_precio else None

                # Skip filas vacías.
                if not val_nombre and not val_codigo and not val_precio:
                    continue

                # ¿Es header de categoría? font.bold + sin código + sin precio.
                es_bold = bool(cell_nombre.font and cell_nombre.font.bold)
                if es_bold and not val_codigo and not val_precio and val_nombre:
                    nombre_cat = _clean_categoria(str(val_nombre))
                    # Resolver rubro desde el mapeo. Match por igualdad
                    # exacta primero (key "MASTICABLE(caramelo)"), si no
                    # por igualdad case-insensitive.
                    rubro_nombre = mapeo.get(nombre_cat)
                    if rubro_nombre is None:
                        for k, v in mapeo.items():
                            if k.lower() == nombre_cat.lower():
                                rubro_nombre = v
                                break
                    if rubro_nombre is None:
                        rubro_nombre = rubro_default
                    rubro = _get_or_create_rubro(rubro_nombre)
                    categoria_actual = _get_or_create_categoria(nombre_cat, rubro)
                    self.stdout.write(self.style.SUCCESS(
                        f'  R{cell_nombre.row:>4} CAT  → {nombre_cat} (rubro: {rubro_nombre})'
                    ))
                    continue

                # Articulo: tiene que tener nombre + precio (código opcional).
                if not val_nombre or val_precio is None:
                    if verbose and val_nombre:
                        self.stdout.write(
                            f'  R{cell_nombre.row:>4} SKIP → {str(val_nombre)[:50]} (sin precio)'
                        )
                    stats['skipped'] += 1
                    continue

                precio = _parse_precio(val_precio)
                if precio is None:
                    if verbose:
                        self.stdout.write(
                            f'  R{cell_nombre.row:>4} SKIP → {str(val_nombre)[:40]} '
                            f'(precio inválido: {val_precio!r})'
                        )
                    stats['skipped'] += 1
                    continue

                if categoria_actual is None and not dry_run:
                    if verbose:
                        self.stdout.write(self.style.WARNING(
                            f'  R{cell_nombre.row:>4} SIN-CAT → {str(val_nombre)[:40]} '
                            f'(no hay categoría activa, saltado)'
                        ))
                    stats['skipped'] += 1
                    continue

                nombre_art = _clean_nombre_articulo(str(val_nombre))
                if not nombre_art or len(nombre_art) < 2:
                    stats['skipped'] += 1
                    continue
                codigo = str(val_codigo).strip() if val_codigo else ''

                # Match: por (codigo + categoria) si tiene código, sino por
                # (nombre + categoria).
                qs = Articulo.objects.all()
                if codigo:
                    qs = qs.filter(codigo=codigo)
                else:
                    qs = qs.filter(nombre__iexact=nombre_art)
                if categoria_actual:
                    qs = qs.filter(categoria=categoria_actual)
                existente = qs.first()

                if existente:
                    cambios = []
                    if update_precios and existente.precio_mayorista != precio:
                        if not dry_run:
                            existente.precio_mayorista = precio
                        cambios.append('precio_mayorista')
                    if existente.nombre != nombre_art:
                        if not dry_run:
                            existente.nombre = nombre_art
                        cambios.append('nombre')
                    if cambios:
                        if not dry_run:
                            existente.save(update_fields=cambios)
                        stats['articulos_actualizados'] += 1
                        if verbose:
                            self.stdout.write(
                                f'  R{cell_nombre.row:>4} UPD  → {nombre_art[:40]} '
                                f'[{", ".join(cambios)}]'
                            )
                    else:
                        stats['articulos_sin_cambios'] += 1
                else:
                    if not dry_run:
                        Articulo.objects.create(
                            codigo=codigo,
                            nombre=nombre_art,
                            precio_mayorista=precio,
                            precio_minorista=precio if minorista_igual else None,
                            categoria=categoria_actual,
                            stock=0,
                            vencimiento=VENCIMIENTO_SENTINEL,
                            marca='Generico',
                        )
                    stats['articulos_creados'] += 1
                    if verbose:
                        self.stdout.write(
                            f'  R{cell_nombre.row:>4} NEW  → {nombre_art[:40]} '
                            f'(cod={codigo or "—"}, ${precio})'
                        )

            if dry_run:
                self.stdout.write(self.style.WARNING(
                    '\n⚠ DRY RUN — abortando, no se guarda nada.'
                ))
                transaction.set_rollback(True)

        # Resumen final.
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('━━━ Resumen ━━━'))
        self.stdout.write(f'  Rubros creados:          {stats["rubros_creados"]}')
        self.stdout.write(f'  Rubros existentes:       {stats["rubros_existentes"]}')
        self.stdout.write(f'  Categorías creadas:      {stats["categorias_creadas"]}')
        self.stdout.write(f'  Categorías existentes:   {stats["categorias_existentes"]}')
        self.stdout.write(f'  Articulos creados:       {stats["articulos_creados"]}')
        self.stdout.write(f'  Articulos actualizados:  {stats["articulos_actualizados"]}')
        self.stdout.write(f'  Articulos sin cambios:   {stats["articulos_sin_cambios"]}')
        self.stdout.write(f'  Filas skipeadas:         {stats["skipped"]}')
